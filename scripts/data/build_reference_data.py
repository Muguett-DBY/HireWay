"""Build the D1 reference catalogue from official open-data releases."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable, Iterator, Sequence

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
# Keep the exact source snapshot beside the scripts that consume it.
RAW_DIR = ROOT / "data" / "sources" / "reference"
GENERATED_DIR = ROOT / "data" / "generated"
SQL_PATH = GENERATED_DIR / "reference_data.sql"
REPORT_PATH = GENERATED_DIR / "reference_report.json"

SOURCES = {
    "osca_descriptions": (
        "OSCA Category Descriptions.xlsx",
        "https://www.abs.gov.au/statistics/classifications/"
        "osca-occupation-standard-classification-australia/2024-version-1-0/"
        "data-downloads/OSCA%20Category%20Descriptions.xlsx",
    ),
    "osca_titles": (
        "OSCA Title Index.xlsx",
        "https://www.abs.gov.au/statistics/classifications/"
        "osca-occupation-standard-classification-australia/2024-version-1-0/"
        "data-downloads/OSCA%20index%20of%20principal%20titles%20alternative%20"
        "titles%20and%20specialisations.xlsx",
    ),
    "osca_crosswalks": (
        "OSCA Correspondence Tables.xlsx",
        "https://www.abs.gov.au/statistics/classifications/"
        "osca-occupation-standard-classification-australia/2024-version-1-0/"
        "data-downloads/OSCA%20correspondence%20tables%20v2.xlsx",
    ),
    "cip_onet": (
        "Education CIP to ONET SOC.xlsx",
        "https://www.onetcenter.org/crosswalks/cip/"
        "Education_CIP_to_ONET_SOC.xlsx",
    ),
    "esco_onet": (
        "ESCO to ONET SOC.xlsx",
        "https://www.onetcenter.org/crosswalks/esco/ESCO_to_ONET-SOC.xlsx",
    ),
    "onet_occupations": (
        "occupation_data.csv",
        "https://www.onetcenter.org/dl_files/database/db_31_0_csv/"
        "occupation_data.csv",
    ),
    "onet_essential": (
        "essential_skills.csv",
        "https://www.onetcenter.org/dl_files/database/db_31_0_csv/"
        "essential_skills.csv",
    ),
    "onet_transferable": (
        "transferable_skills.csv",
        "https://www.onetcenter.org/dl_files/database/db_31_0_csv/"
        "transferable_skills.csv",
    ),
    "onet_software": (
        "software_skills.csv",
        "https://www.onetcenter.org/dl_files/database/db_31_0_csv/"
        "software_skills.csv",
    ),
}


@dataclass(frozen=True)
class SkillRecord:
    code: str
    name: str
    description: str
    kind: str
    source: str


def download_sources() -> dict[str, Path]:
    """Download missing source files and keep them out of Git."""

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}

    for key, (filename, url) in SOURCES.items():
        path = RAW_DIR / filename
        paths[key] = path
        if path.exists() and path.stat().st_size > 0:
            continue

        print(f"Downloading {filename}...")
        request = urllib.request.Request(
            url,
            headers={"User-Agent": "HireWay data import (student project)"},
        )
        with urllib.request.urlopen(request, timeout=120) as response:
            path.write_bytes(response.read())

    return paths


def clean_text(value: object) -> str:
    """Turn spreadsheet cells into compact display text."""

    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_code(value: object) -> str:
    """Keep classification codes as text, including leading zeroes."""

    text = clean_text(value)
    return text[:-2] if text.endswith(".0") else text


def normalise_name(value: str) -> str:
    """Create a stable key for case and punctuation-insensitive matching."""

    value = value.casefold().replace("&", " and ")
    return " ".join(re.findall(r"[a-z0-9+#.]+", value))


def title_tokens(value: str) -> set[str]:
    """Reduce small title differences before comparing occupations."""

    ignored = {"and", "of", "the", "other", "general", "all"}
    tokens: set[str] = set()
    for token in re.findall(r"[a-z0-9]+", value.casefold()):
        if token in ignored:
            continue
        if token.endswith("ies") and len(token) > 4:
            token = f"{token[:-3]}y"
        elif token.endswith(("ches", "shes", "xes")) and len(token) > 4:
            token = token[:-2]
        elif token.endswith("s") and not token.endswith("ss") and len(token) > 3:
            token = token[:-1]
        tokens.add(token)
    return tokens


def title_score(left: str, right: str) -> float:
    """Score two occupation titles after the official code narrows the search."""

    left_key = " ".join(sorted(title_tokens(left)))
    right_key = " ".join(sorted(title_tokens(right)))
    if not left_key or not right_key:
        return 0.0
    if left_key == right_key:
        return 1.0

    left_tokens = set(left_key.split())
    right_tokens = set(right_key.split())
    overlap = len(left_tokens & right_tokens) / len(left_tokens | right_tokens)
    sequence = SequenceMatcher(None, left_key, right_key).ratio()
    return round((0.65 * overlap) + (0.35 * sequence), 4)


def workbook_rows(path: Path, sheet: str, start_row: int) -> Iterator[tuple]:
    """Read an Excel sheet without loading the whole workbook into memory."""

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    try:
        yield from worksheet.iter_rows(min_row=start_row, values_only=True)
    finally:
        workbook.close()


def read_osca(paths: dict[str, Path]) -> tuple[dict, set, dict]:
    """Read OSCA occupations, searchable titles, and ISCO links."""

    occupations: dict[str, dict[str, object]] = {}
    for row in workbook_rows(paths["osca_descriptions"], "Table 1", 6):
        code = clean_code(row[0])
        if not re.fullmatch(r"\d{6}", code):
            continue
        occupations[code] = {
            "title": clean_text(row[1]),
            "description": clean_text(row[3]),
            "names": set(),
        }

    aliases: set[tuple[str, str]] = set()
    for row in workbook_rows(paths["osca_titles"], "Table 1", 6):
        code = clean_code(row[0])
        title = clean_text(row[1])
        category = clean_text(row[2]).casefold()
        if code not in occupations or not title:
            continue
        occupations[code]["names"].add(title)
        if "principal" not in category and title != occupations[code]["title"]:
            aliases.add((code, title))

    osca_to_isco: dict[str, set[str]] = defaultdict(set)
    current_osca = ""
    for row in workbook_rows(paths["osca_crosswalks"], "Table 8", 6):
        row_code = clean_code(row[0])
        if row_code:
            current_osca = row_code
        isco_code = clean_code(row[2]).zfill(4)
        if current_osca in occupations and re.fullmatch(r"\d{4}", isco_code):
            osca_to_isco[current_osca].add(isco_code)

    return occupations, aliases, osca_to_isco


def read_onet_occupations(path: Path) -> dict[str, tuple[str, str]]:
    """Read the current O*NET occupation catalogue."""

    occupations: dict[str, tuple[str, str]] = {}
    with path.open(encoding="utf-8-sig", newline="") as source:
        for row in csv.DictReader(source):
            code = clean_text(row["O*NET-SOC Code"])
            occupations[code] = (
                clean_text(row["Title"]),
                clean_text(row["Description"]),
            )
    return occupations


def read_cip_crosswalk(path: Path) -> tuple[dict, set]:
    """Read fields of study and their official O*NET links."""

    programs: dict[str, str] = {}
    mappings: set[tuple[str, str]] = set()
    for row in workbook_rows(path, "O-NET-SOC 2019 Crosswalks", 4):
        cip_code = clean_code(row[0])
        title = clean_text(row[1]).removesuffix(".")
        onet_code = clean_text(row[2])
        if not cip_code or not title or not onet_code:
            continue
        code = f"CIP:{cip_code}"
        programs[code] = title
        mappings.add((code, onet_code))
    return programs, mappings


def read_esco_crosswalk(path: Path) -> dict[str, list[tuple[str, str, str]]]:
    """Group the official ESCO/O*NET crosswalk by four-digit ISCO code."""

    by_isco: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for row in workbook_rows(path, "O-NET-SOC 2019 Crosswalks", 5):
        esco_code = clean_text(row[0])
        esco_title = clean_text(row[1])
        onet_code = clean_text(row[2])
        onet_title = clean_text(row[3])
        isco_code = esco_code[:4]
        if re.fullmatch(r"\d{4}", isco_code) and onet_code:
            by_isco[isco_code].append((esco_title, onet_code, onet_title))
    return by_isco


def build_occupation_links(
    occupations: dict,
    osca_to_isco: dict[str, set[str]],
    esco_by_isco: dict[str, list[tuple[str, str, str]]],
    onet_occupations: dict[str, tuple[str, str]],
) -> set[tuple[str, str, float, str]]:
    """Keep only strong title matches inside an official ISCO grouping."""

    links: set[tuple[str, str, float, str]] = set()
    for osca_code, occupation in occupations.items():
        names = set(occupation["names"]) | {str(occupation["title"])}
        scores: dict[str, float] = defaultdict(float)

        for isco_code in osca_to_isco.get(osca_code, set()):
            for esco_title, onet_code, crosswalk_title in esco_by_isco.get(
                isco_code, []
            ):
                if onet_code not in onet_occupations:
                    continue
                candidate_titles = {
                    esco_title,
                    crosswalk_title,
                    onet_occupations[onet_code][0],
                }
                score = max(
                    title_score(name, candidate)
                    for name in names
                    for candidate in candidate_titles
                    if candidate
                )
                scores[onet_code] = max(scores[onet_code], score)

        if not scores:
            continue
        best_score = max(scores.values())
        if best_score < 0.78:
            continue

        for onet_code, score in sorted(
            scores.items(), key=lambda item: (-item[1], item[0])
        )[:3]:
            if score >= 0.78 and score >= best_score - 0.04:
                links.add((osca_code, onet_code, score, "isco_title_match"))

    return links


def friendly_tool_name(name: str) -> str:
    """Use familiar labels for a few verbose O*NET software names."""

    replacements = {
        "Structured query language SQL": "SQL",
        "Microsoft Power BI": "Power BI",
        "Microsoft Visual Studio Code": "Visual Studio Code",
    }
    return replacements.get(name, name)


def make_tool_code(name: str) -> str:
    """Create a stable code because O*NET software examples have no own ID."""

    digest = hashlib.sha256(normalise_name(name).encode()).hexdigest()[:16]
    return f"onet-tool:{digest}"


def read_onet_skills(paths: dict[str, Path]) -> tuple[dict, set, set]:
    """Combine O*NET skills and named software into one searchable catalogue."""

    skills_by_name: dict[str, SkillRecord] = {}
    aliases: set[tuple[str, str]] = set()
    relationships: dict[tuple[str, str], tuple[float, int, int, str]] = {}

    # Importance is reported on a zero-to-five scale, so convert it to percent.
    for key, requirement_type in (
        ("onet_essential", "essential_skill"),
        ("onet_transferable", "transferable_skill"),
    ):
        with paths[key].open(encoding="utf-8-sig", newline="") as source:
            for row in csv.DictReader(source):
                if row["Scale ID"] != "IM" or row["Recommend Suppress"] == "Y":
                    continue
                name = clean_text(row["Element Name"])
                name_key = normalise_name(name)
                suggested_code = f"onet-skill:{clean_text(row['Element ID'])}"
                record = skills_by_name.setdefault(
                    name_key,
                    SkillRecord(suggested_code, name, "", "skill", "O*NET 31.0"),
                )
                code = record.code
                score = min(100.0, float(row["Data Value"]) * 20)
                relationship_key = (clean_text(row["O*NET-SOC Code"]), code)
                old = relationships.get(
                    relationship_key,
                    (0.0, 0, 0, requirement_type),
                )
                relationships[relationship_key] = (
                    max(score, old[0]),
                    0,
                    0,
                    requirement_type,
                )

    # Named workplace examples make the recommendations useful to real users.
    with paths["onet_software"].open(
        encoding="utf-8-sig", newline=""
    ) as source:
        for row in csv.DictReader(source):
            source_name = clean_text(row["Workplace Example"])
            name = friendly_tool_name(source_name)
            name_key = normalise_name(name)
            if name_key in skills_by_name:
                code = skills_by_name[name_key].code
            else:
                code = make_tool_code(name)
                skills_by_name[name_key] = SkillRecord(
                    code,
                    name,
                    clean_text(row["Element Name"]),
                    "tool",
                    "O*NET 31.0",
                )
            if name != source_name:
                aliases.add((code, source_name))

            hot = int(row["Hot Technology"] == "Y")
            in_demand = int(row["In Demand"] == "Y")
            score = 55 + (25 * hot) + (20 * in_demand)
            relationship_key = (clean_text(row["O*NET-SOC Code"]), code)
            old = relationships.get(relationship_key, (0.0, 0, 0, "tool"))
            relationships[relationship_key] = (
                max(float(score), old[0]),
                max(hot, old[1]),
                max(in_demand, old[2]),
                "tool",
            )

    return skills_by_name, aliases, {
        (onet_code, skill_code, score, hot, in_demand, requirement_type)
        for (onet_code, skill_code), (
            score,
            hot,
            in_demand,
            requirement_type,
        ) in relationships.items()
    }


def sql_value(value: object) -> str:
    """Encode generated values as SQLite literals."""

    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def batched(items: Sequence[tuple], size: int = 150) -> Iterator[Sequence[tuple]]:
    """Keep each generated SQL statement small enough for Wrangler and D1."""

    for start in range(0, len(items), size):
        yield items[start : start + size]


def insert_many(
    table: str,
    columns: Sequence[str],
    rows: Iterable[tuple],
    conflict_sql: str = "DO NOTHING",
) -> list[str]:
    """Create compact multi-row upserts for the generated SQL file."""

    ordered_rows = sorted(set(rows), key=lambda row: tuple(str(value) for value in row))
    statements: list[str] = []
    column_sql = ", ".join(columns)
    for batch in batched(ordered_rows):
        values = ",\n  ".join(
            "(" + ", ".join(sql_value(value) for value in row) + ")"
            for row in batch
        )
        statements.append(
            f"INSERT INTO {table} ({column_sql}) VALUES\n  {values}\n"
            f"ON CONFLICT {conflict_sql};"
        )
    return statements


def sha256(path: Path) -> str:
    """Record the exact downloaded release used for this build."""

    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_reference_data(
    occupations: dict,
    occupation_aliases: set[tuple[str, str]],
    onet_occupations: dict[str, tuple[str, str]],
    programs: dict[str, str],
    education_links: set[tuple[str, str]],
    skills_by_name: dict[str, SkillRecord],
    skill_aliases: set[tuple[str, str]],
    occupation_skills: set[tuple[str, str, float, int, int]],
    occupation_links: set[tuple[str, str, float, str]],
) -> dict[str, object]:
    """Stop the build when a stable key or relationship is broken."""

    skill_codes = {record.code for record in skills_by_name.values()}
    failures: list[str] = []

    checks = {
        "occupation_alias_parents": all(
            code in occupations for code, _ in occupation_aliases
        ),
        "education_link_parents": all(
            education_code in programs and onet_code in onet_occupations
            for education_code, onet_code in education_links
        ),
        "skill_alias_parents": all(
            skill_code in skill_codes for skill_code, _ in skill_aliases
        ),
        "occupation_skill_parents": all(
            onet_code in onet_occupations and skill_code in skill_codes
            for onet_code, skill_code, *_ in occupation_skills
        ),
        "occupation_link_parents": all(
            occupation_code in occupations and onet_code in onet_occupations
            for occupation_code, onet_code, *_ in occupation_links
        ),
        "skill_scores_in_range": all(
            0 <= score <= 100 and hot in (0, 1) and in_demand in (0, 1)
            for _, _, score, hot, in_demand, _ in occupation_skills
        ),
        "requirement_types_known": all(
            requirement_type
            in {"essential_skill", "transferable_skill", "tool"}
            for *_, requirement_type in occupation_skills
        ),
        "mapping_confidence_in_range": all(
            0.78 <= confidence <= 1
            for _, _, confidence, _ in occupation_links
        ),
        "normalised_skill_names_unique": len(skills_by_name) == len(skill_codes),
        "data_science_program_present": programs.get("CIP:30.7001")
        == "Data Science, General",
        "data_scientist_goal_present": occupations.get("223234", {}).get("title")
        == "Data Scientist",
        "data_scientist_mapping_present": any(
            occupation_code == "223234" and onet_code == "15-2051.00"
            for occupation_code, onet_code, *_ in occupation_links
        ),
        "python_present": "python" in skills_by_name,
        "sql_present": "sql" in skills_by_name,
    }

    for name, passed in checks.items():
        if not passed:
            failures.append(name)
    if failures:
        raise ValueError("Reference data checks failed: " + ", ".join(failures))

    return {"status": "passed", "checks": checks}


def build_sql(paths: dict[str, Path]) -> dict[str, object]:
    """Parse all sources and write one repeatable D1 import file."""

    occupations, occupation_aliases, osca_to_isco = read_osca(paths)
    onet_occupations = read_onet_occupations(paths["onet_occupations"])
    programs, education_links = read_cip_crosswalk(paths["cip_onet"])
    esco_by_isco = read_esco_crosswalk(paths["esco_onet"])
    occupation_links = build_occupation_links(
        occupations, osca_to_isco, esco_by_isco, onet_occupations
    )
    skills_by_name, skill_aliases, occupation_skills = read_onet_skills(paths)

    # Drop links whose O*NET code is not part of the current release.
    education_links = {
        link for link in education_links if link[1] in onet_occupations
    }
    occupation_links = {
        link for link in occupation_links if link[1] in onet_occupations
    }
    occupation_skills = {
        link for link in occupation_skills if link[0] in onet_occupations
    }

    quality = validate_reference_data(
        occupations,
        occupation_aliases,
        onet_occupations,
        programs,
        education_links,
        skills_by_name,
        skill_aliases,
        occupation_skills,
        occupation_links,
    )

    source_rows = [
        (
            "ABS OSCA 2024",
            "Australian Bureau of Statistics",
            SOURCES["osca_descriptions"][1],
            "Creative Commons Attribution 4.0 International",
            "Australian occupation titles, aliases and descriptions.",
            date.today().isoformat(),
        ),
        (
            "O*NET 31.0",
            "U.S. Department of Labor, Employment and Training Administration",
            "https://www.onetcenter.org/database.html",
            "Creative Commons Attribution 4.0 International",
            "Occupation, skill, technology, CIP and ESCO crosswalk data.",
            date.today().isoformat(),
        ),
    ]

    statements = [
        "-- Generated by scripts/data/build_reference_data.py.",
        "-- User profiles are never deleted or replaced by this import.",
        "PRAGMA foreign_keys = ON;",
    ]
    statements += insert_many(
        "data_source",
        ("name", "publisher", "source_url", "licence", "description", "accessed_on"),
        source_rows,
        "(name) DO UPDATE SET publisher = excluded.publisher, "
        "source_url = excluded.source_url, licence = excluded.licence, "
        "description = excluded.description, accessed_on = excluded.accessed_on",
    )

    release_rows = []
    for key, (filename, _) in SOURCES.items():
        source_name = "ABS OSCA 2024" if key.startswith("osca_") else "O*NET 31.0"
        release_rows.append((source_name, filename, sha256(paths[key])))
    for source_name, filename, checksum in release_rows:
        statements.append(
            "INSERT INTO dataset_release "
            "(data_source_id, release_label, published_on, source_file, checksum_sha256) "
            f"SELECT id, {sql_value('2024 v1.0' if source_name.startswith('ABS') else '31.0')}, "
            f"NULL, {sql_value(filename)}, {sql_value(checksum)} FROM data_source "
            f"WHERE name = {sql_value(source_name)} "
            "ON CONFLICT (data_source_id, release_label, source_file) "
            "DO UPDATE SET checksum_sha256 = excluded.checksum_sha256;"
        )

    statements += insert_many(
        "occupation",
        ("code", "title", "description"),
        (
            (code, values["title"], values["description"])
            for code, values in occupations.items()
        ),
        "(code) DO UPDATE SET title = excluded.title, "
        "description = excluded.description, updated_at = CURRENT_TIMESTAMP",
    )
    statements.append("DELETE FROM occupation_alias;")
    statements += insert_many(
        "occupation_alias",
        ("occupation_code", "alias"),
        occupation_aliases,
        "(occupation_code, alias) DO NOTHING",
    )
    statements += insert_many(
        "education_program",
        ("code", "title", "source"),
        ((code, title, "CIP 2020") for code, title in programs.items()),
        "(code) DO UPDATE SET title = excluded.title, source = excluded.source",
    )
    statements += insert_many(
        "onet_occupation",
        ("code", "title", "description"),
        (
            (code, title, description)
            for code, (title, description) in onet_occupations.items()
        ),
        "(code) DO UPDATE SET title = excluded.title, description = excluded.description",
    )
    statements += insert_many(
        "skill",
        ("code", "name", "description", "kind", "source"),
        (
            (record.code, record.name, record.description, record.kind, record.source)
            for record in skills_by_name.values()
        ),
        "(code) DO UPDATE SET name = excluded.name, "
        "description = excluded.description, kind = excluded.kind, source = excluded.source",
    )
    statements.append("DELETE FROM skill_alias;")
    statements += insert_many(
        "skill_alias",
        ("skill_code", "alias"),
        skill_aliases,
        "(skill_code, alias) DO NOTHING",
    )

    # Rebuild only generated bridge tables; user-owned profile rows stay untouched.
    statements += [
        "DELETE FROM education_onet_map;",
        "DELETE FROM onet_occupation_skill;",
        "DELETE FROM occupation_onet_map;",
    ]
    statements += insert_many(
        "education_onet_map",
        ("education_code", "onet_code"),
        education_links,
        "(education_code, onet_code) DO NOTHING",
    )
    statements += insert_many(
        "onet_occupation_skill",
        (
            "onet_code",
            "skill_code",
            "score",
            "hot_technology",
            "in_demand",
            "requirement_type",
        ),
        occupation_skills,
        "(onet_code, skill_code) DO UPDATE SET score = excluded.score, "
        "hot_technology = excluded.hot_technology, in_demand = excluded.in_demand, "
        "requirement_type = excluded.requirement_type",
    )
    statements += insert_many(
        "occupation_onet_map",
        ("occupation_code", "onet_code", "confidence", "method"),
        occupation_links,
        "(occupation_code, onet_code) DO UPDATE SET confidence = excluded.confidence, "
        "method = excluded.method",
    )
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    SQL_PATH.write_text("\n\n".join(statements) + "\n", encoding="utf-8")

    mapped_goals = {link[0] for link in occupation_links}
    report = {
        "files": {
            key: {"filename": path.name, "sha256": sha256(path)}
            for key, path in paths.items()
        },
        "counts": {
            "education_programs": len(programs),
            "education_onet_links": len(education_links),
            "osca_occupations": len(occupations),
            "occupation_aliases": len(occupation_aliases),
            "mapped_osca_occupations": len(mapped_goals),
            "occupation_onet_links": len(occupation_links),
            "onet_occupations": len(onet_occupations),
            "skills_and_tools": len(skills_by_name),
            "onet_occupation_skill_links": len(occupation_skills),
        },
        "quality": quality,
        "samples": {
            "data_programs": sorted(
                (code, title)
                for code, title in programs.items()
                if "data" in title.casefold()
            )[:12],
            "data_goals": sorted(
                (code, values["title"])
                for code, values in occupations.items()
                if "data" in str(values["title"]).casefold()
            )[:12],
            "data_goal_links": sorted(
                link
                for link in occupation_links
                if "data" in str(occupations[link[0]]["title"]).casefold()
            )[:12],
        },
    }
    REPORT_PATH.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return report


def main() -> None:
    """Download, transform, and report the generated reference data."""

    paths = download_sources()
    report = build_sql(paths)
    print(json.dumps(report["counts"], indent=2))
    print(f"Wrote {SQL_PATH.relative_to(ROOT)}")
    print(f"Wrote {REPORT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
